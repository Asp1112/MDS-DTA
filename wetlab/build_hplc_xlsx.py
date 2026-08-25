"""Extract HPLC-related tables from the wet-lab supplementary document."""
import re
from pathlib import Path

import docx
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = Path(r"第一轮审稿湿实验补充数据.docx")
OUT = Path("HPLC_data.xlsx")
KEYWORDS = ("HPLC", "hplc", "LOD", "LOQ", "retention", "peak area",
            "quantification", "validation", "p-aminophenol", "acetaminophen",
            "standard curve", "linear", "recovery")


def table_heading(text: str) -> str:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    return " | ".join(lines[:3])[:160]


def main() -> None:
    doc = docx.Document(SRC)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    n = 0
    for ti, table in enumerate(doc.tables, 1):
        rows = [[(c.text or "").strip() for c in row.cells] for row in table.rows]
        flat = " ".join(" ".join(r) for r in rows)
        if not re.search("|".join(KEYWORDS), flat, re.IGNORECASE):
            continue
        n += 1
        ws = wb.create_sheet(f"Table_{ti}")
        for r, row in enumerate(rows, 1):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                if r == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        for col in range(1, max((len(r) for r in rows), default=1) + 1):
            width = min(60, max((len(r[col - 1]) for r in rows if len(r) >= col), default=8) + 2)
            ws.column_dimensions[get_column_letter(col)].width = width
    wb.save(OUT)
    print(f"extracted {n} HPLC-related tables -> {OUT}")


if __name__ == "__main__":
    main()
